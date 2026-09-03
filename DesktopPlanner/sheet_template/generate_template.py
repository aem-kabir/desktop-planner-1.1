"""
generate_template.py
Генерирует Tasks_Template.xlsx — шаблон для импорта в Google Таблицы:
- заголовки столбцов ровно как ожидает приложение (сопоставление по имени),
- 2-3 строки-примера,
- выпадающие списки (data validation) для Status и Priority на 500 строк вперёд,
- предзаполненные дефолты (Active/Medium) в пустых строках-заготовках.
"""
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

HEADERS = [
    "ID", "Title", "Description", "StartDate", "DeadlineDate",
    "DeadlineTime", "Status", "Priority", "Assignee", "UpdatedAt",
]

STATUS_VALUES = ["Active", "Completed", "Canceled"]
PRIORITY_VALUES = ["High", "Medium", "Low"]

TOTAL_ROWS = 500          # диапазон для выпадающих списков и заготовок
BLANK_ROWS_WITH_DEFAULTS = 40  # сколько пустых строк-заготовок предзаполнить дефолтами

HEADER_FILL = PatternFill(start_color="3D7EFF", end_color="3D7EFF", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Arial", size=11)
EXAMPLE_FILL = PatternFill(start_color="EAF1FF", end_color="EAF1FF", fill_type="solid")
DEFAULT_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"

    # --- Заголовки ---
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    # --- Ширины столбцов ---
    widths = {
        "A": 24,  # ID
        "B": 28,  # Title
        "C": 40,  # Description
        "D": 12,  # StartDate
        "E": 14,  # DeadlineDate
        "F": 13,  # DeadlineTime
        "G": 11,  # Status
        "H": 10,  # Priority
        "I": 16,  # Assignee
        "J": 20,  # UpdatedAt
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # --- Примеры (2-3 строки) ---
    now_iso = datetime.datetime.now().replace(microsecond=0).isoformat()
    today = datetime.date.today()
    tomorrow = today + datetime.timedelta(days=1)
    next_week = today + datetime.timedelta(days=7)

    examples = [
        [
            "11111111-1111-1111-1111-111111111111",
            "Пример: подготовить отчёт",
            "Собрать данные за месяц и оформить в презентацию",
            today.isoformat(), tomorrow.isoformat(), "18:00",
            "Active", "High", "Иван Иванов", now_iso,
        ],
        [
            "22222222-2222-2222-2222-222222222222",
            "Пример: позвонить клиенту",
            "",
            today.isoformat(), today.isoformat(), "15:30",
            "Active", "Medium", "", now_iso,
        ],
        [
            "33333333-3333-3333-3333-333333333333",
            "Пример: сдать квартальный отчёт",
            "Финальная проверка перед отправкой руководству",
            today.isoformat(), next_week.isoformat(), "",
            "Active", "Low", "Отдел финансов", now_iso,
        ],
    ]

    row_num = 2
    for example in examples:
        for col_idx, value in enumerate(example, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = DEFAULT_FONT
            cell.fill = EXAMPLE_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 3))
        row_num += 1

    # --- Пустые строки-заготовки с предзаполненными дефолтами ---
    # Чтобы человеку, добавляющему задачу вручную, не нужно было печатать
    # Status/Priority и нельзя было опечататься — они уже стоят как Active/Medium.
    first_blank_row = row_num
    last_blank_row = first_blank_row + BLANK_ROWS_WITH_DEFAULTS - 1
    for r in range(first_blank_row, last_blank_row + 1):
        # ID оставляем пустым намеренно: приложение генерирует UUID на клиенте.
        # Если человек добавляет задачу вручную прямо в таблице и оставляет ID
        # пустым, при следующей синхронизации виджет присвоит ей новый UUID.
        status_cell = ws.cell(row=r, column=HEADERS.index("Status") + 1, value="Active")
        priority_cell = ws.cell(row=r, column=HEADERS.index("Priority") + 1, value="Medium")
        for c in (status_cell, priority_cell):
            c.font = DEFAULT_FONT
            c.border = THIN_BORDER
        for col_idx in range(1, len(HEADERS) + 1):
            ws.cell(row=r, column=col_idx).border = THIN_BORDER

    # --- Data validation (выпадающие списки) на TOTAL_ROWS строк вперёд ---
    status_col_letter = get_column_letter(HEADERS.index("Status") + 1)
    priority_col_letter = get_column_letter(HEADERS.index("Priority") + 1)

    status_dv = DataValidation(
        type="list",
        formula1=f'"{",".join(STATUS_VALUES)}"',
        allow_blank=True,
        showDropDown=False,  # False в openpyxl фактически ПОКАЗЫВАЕТ стрелку списка
    )
    status_dv.error = "Допустимые значения: Active, Completed, Canceled"
    status_dv.errorTitle = "Некорректный статус"
    status_dv.prompt = "Выберите статус задачи из списка"
    status_dv.promptTitle = "Status"
    ws.add_data_validation(status_dv)
    status_dv.add(f"{status_col_letter}2:{status_col_letter}{TOTAL_ROWS}")

    priority_dv = DataValidation(
        type="list",
        formula1=f'"{",".join(PRIORITY_VALUES)}"',
        allow_blank=True,
        showDropDown=False,
    )
    priority_dv.error = "Допустимые значения: High, Medium, Low"
    priority_dv.errorTitle = "Некорректный приоритет"
    priority_dv.prompt = "Выберите приоритет задачи из списка"
    priority_dv.promptTitle = "Priority"
    ws.add_data_validation(priority_dv)
    priority_dv.add(f"{priority_col_letter}2:{priority_col_letter}{TOTAL_ROWS}")

    # --- Второй лист: краткая инструкция/легенда ---
    legend = wb.create_sheet("Инструкция")
    legend.column_dimensions["A"].width = 100
    legend_lines = [
        ("Как пользоваться этой таблицей", True),
        ("", False),
        ("1. Это шаблон для листа 'Tasks', с которым синхронизируется desktop-виджет «Планер задач».", False),
        ("2. Названия столбцов на листе Tasks менять НЕЛЬЗЯ — приложение ищет их по имени.", False),
        ("3. Столбец ID — не трогайте вручную для существующих задач. Для НОВОЙ задачи, добавляемой", False),
        ("   прямо в таблице, можно оставить ID пустым — виджет присвоит уникальный идентификатор", False),
        ("   при следующей синхронизации.", False),
        ("4. Status и Priority — выбирайте из выпадающего списка (щёлкните по ячейке, появится стрелка).", False),
        ("   Если оставить их пустыми, виджет применит значения по умолчанию: Active и Medium.", False),
        ("5. DeadlineTime можно оставить пустым — тогда дедлайн считается на конец дня, 23:59.", False),
        ("6. Формат дат: ГГГГ-ММ-ДД (например 2026-09-15). Формат времени: ЧЧ:ММ (например 18:30).", False),
        ("7. UpdatedAt заполняется автоматически виджетом и используется для разрешения конфликтов", False),
        ("   при синхронизации (побеждает более свежая по времени запись). Менять его вручную не нужно.", False),
        ("", False),
        ("Как подключить эту таблицу к виджету — см. файл SETUP_GOOGLE_SHEETS.md в архиве проекта.", True),
    ]
    for i, (text, bold) in enumerate(legend_lines, start=1):
        cell = legend.cell(row=i, column=1, value=text)
        cell.font = Font(name="Arial", size=11, bold=bold)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    return wb


if __name__ == "__main__":
    workbook = build_workbook()
    out_path = "Tasks_Template.xlsx"
    workbook.save(out_path)
    print(f"Шаблон сохранён: {out_path}")
